import openpyxl
import os
import shutil
import sys
import logging
import datetime
import tkinter as tk
from tkinter import filedialog, messagebox


# 获取当前脚本所在目录的父目录路径
parent_folder_path = os.path.dirname(os.path.dirname(__file__))
print(f'parent_folder_path{parent_folder_path}')
# 拼接数据文件夹的路径
oe_path = os.path.dirname(parent_folder_path)
print(f"oe_path{oe_path}")
# 拼接 Excel 模板文件的完整路径
OE_EXCEL_TEMPLATE_DIR = oe_path + "/data/import_excel_template_27772.xlsx"


# 从指定的 Dcase 测试用例 Excel 文件中提取数据
def getDcaseData(dcaseDir):
    try:
        # 加载 Excel 文件
        wb = openpyxl.load_workbook(dcaseDir)
        # 获取第一个工作表
        sheet = wb.worksheets[0]
        # 初始化测试用例数据起始行
        caseStartRow = 0
        # 初始化各列索引
        caseNameCol = 0
        caseLevelCol = 0
        caseStepCol = 0
        caseResultCol = 0
        casePreconditionCol = 0
        # 定义需要查找的表头元素
        elements = ("caseId", "名称", "级别", "步骤", "预期结果")
        # 遍历工作表的每一行
        for row_index, row in enumerate(sheet.iter_rows(min_row=1, max_row=sheet.max_row, values_only=True), start=1):
            flag = 0
            # 检查当前行是否包含所有需要的表头元素
            for element in elements:
                if any(x == element for x in row):
                    flag += 1
                    # 记录各表头元素所在的列索引
                    if element == "名称":
                        caseNameCol = row.index(element)
                    if element == "级别":
                        caseLevelCol = row.index(element)
                    if element == "步骤":
                        caseStepCol = row.index(element)
                    if element == "预期结果":
                        caseResultCol = row.index(element)
                    if element == "前提条件":
                        casePreconditionCol = row.index(element)
            # 若找到包含所有表头元素的行，记录该行索引并跳出循环
            if flag == 5:
                caseStartRow = row_index
                break
        dcaseContent = []
        # 从表头下一行开始遍历，提取测试用例数据
        for row in sheet.iter_rows(min_row=caseStartRow + 1, max_row=sheet.max_row, values_only=True):
            dcaseContent.append((row[caseNameCol], row[caseStepCol],
                                 row[caseResultCol], row[caseLevelCol],
                                 row[casePreconditionCol]))
        # 关闭 Excel 文件
        wb.close()
        return dcaseContent
    except Exception as e:
        # 若出现异常，记录错误信息
        logging.error(f"解析 Dcase 测试用例文件 {dcaseDir} 时出错: {e}")
        return []


# 格式化列表内容为指定格式的字符串
def formatListContent(tempList, keyStr):
    tempStr = ""
    for i in range(len(tempList)):
        tempStr += f"{keyStr}{i + 1}:{tempList[i]}\n"
    return tempStr


# 删除列表中的空元素和空列表
def deleteSpace(tempList):
    newList = [x for x in tempList if x is not None and str(x).strip() != "" and not (isinstance(x, list) and not x)]
    return newList


# 解析从 Dcase 文件中提取的测试用例数据
def parseCase(dcaseContent):
    tempConent = []
    for index, one in enumerate(dcaseContent):
        try:
            # 格式化步骤和预期结果
            tempStep = formatListContent(deleteSpace(one[1].split("\n")), "step")
            tempResult = formatListContent(deleteSpace(one[2].split("\n")), "assert")
            # 替换级别标识
            tempLevel = one[3].replace("D", "p")
            tempConent.append((one[0], tempStep, tempResult, tempLevel, one[4]))
        except Exception as e:
            # 若出现异常，记录错误信息
            logging.error(f"处理第 {index + 1} 条测试用例时出错: {e}")
    return tempConent


# 根据解析后的 Dcase 数据创建 OE 测试用例文件
def createOECase(dcaseContent, output_path=None):
    ctime = create_format_time()
    out_xlse = ctime + ".xlsx"
    try:
        # 拼接输出文件的路径
        out_xlsx_path = oe_path + "\\" + "output"
        # out_xlsx_path = "/output"
        if not os.path.exists(out_xlsx_path):
            os.makedirs(out_xlsx_path)
        outCaseDir = os.path.join(out_xlsx_path, out_xlse)
        # 复制模板文件到输出路径
        shutil.copyfile(OE_EXCEL_TEMPLATE_DIR, outCaseDir)
        # 加载输出文件
        wb = openpyxl.load_workbook(outCaseDir)
        # 获取指定工作表
        sheet = wb["Data"]
        # 初始化各列索引
        oeCaseNameCol = 0
        oeCaseLevelCol = 0
        oeCaseStepCol = 0
        oeCaseResultCol = 0
        oeCasePreconditionCol = 0
        # 查找表头各元素所在的列索引
        for cell in sheet[1]:
            if cell.value == "用例名":
                oeCaseNameCol = cell.column
            if cell.value == "优先级":
                oeCaseLevelCol = cell.column
            if cell.value == "步骤":
                oeCaseStepCol = cell.column
            if cell.value == "校验点":
                oeCaseResultCol = cell.column
            if cell.value == "前置条件":
                oeCasePreconditionCol = cell.column
        # 将解析后的测试用例数据写入输出文件
        for i in range(len(dcaseContent)):
            oneCase = dcaseContent[i]
            sheet.cell(row=i + 2, column=oeCaseNameCol, value=oneCase[0])
            sheet.cell(row=i + 2, column=oeCaseStepCol, value=oneCase[1])
            sheet.cell(row=i + 2, column=oeCaseResultCol, value=oneCase[2])
            sheet.cell(row=i + 2, column=oeCaseLevelCol, value=oneCase[3])
            sheet.cell(row=i + 2, column=oeCasePreconditionCol, value=oneCase[4])
        # 保存输出文件
        wb.save(outCaseDir)
        messagebox.showinfo(message="转换完成\n文件保存路径\n" + oe_path + "\output\\" + out_xlse)
    except Exception as e:
        # 若出现异常，记录错误信息
        logging.error(f"创建 OE 测试用例文件时出错: {e}")

def create_format_time():
    format_time = datetime.datetime.now().strftime("%Y_%m_%d_%H_%M_%S")
    return format_time



# 选择 Dcase 测试用例文件并进行处理
def select_file():
    # 打开文件选择对话框，选择 Excel 文件
    file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
    try:
        if file_path:
            # 提取 Dcase 文件中的数据
            dcaseContent = getDcaseData(file_path)
            # 创建 OE 测试用例文件
            createOECase(parseCase(dcaseContent))
        else:
            messagebox.showinfo(message="未选择文件")
    except Exception as e:
        logging.error(f"转换失败{e}")

if __name__ == "__main__":
    # 创建主窗口
    root = tk.Tk()
    # 设置窗口标题
    root.title("Excel 测试用例处理工具")

    # 创建选择文件按钮，点击时调用 select_file 函数
    select_button = tk.Button(root, text="选择 Dcase 测试用例文件", command=select_file)
    select_button.pack(pady=20)

    # 创建状态标签，用于显示处理状态
    status_label = tk.Label(root, text="请选择文件...")
    status_label.pack(pady=10)

    # 运行主循环，使窗口保持显示
    root.mainloop()
