import atexit

# import coverage

# cov = coverage.Coverage(branch=True, concurrency="thread", config_file=".coveragerc")
# cov.start()
import openpyxl
import os
import shutil
import sys
import logging

CURRENT_DIR = os.path.split(os.path.realpath("__file__"))[0]
OE_EXCEL_TEMPLATE_DIR = os.path.join(CURRENT_DIR, "import_excel_template_27772.xlsx")


def getDcaseData(dcaseDir):
    """
    通过指定dcase测试用例的地址，解析出包含名称、级别、步骤、预期结果、前提条件的list
    :param dcaseDir:
    :return:
    """
    wb = openpyxl.load_workbook(dcaseDir)
    # sheet = wb['Dcase平台用例导入模版']  # 通过名称选择工作表
    sheet = wb.worksheets[0]  # 默认获取第一个sheet页
    caseStartRow = 0
    caseNameCol = 0
    caseLevelCol = 0
    caseStepCol = 0
    caseResultCol = 0
    casePreconditionCol = 0
    elements = ("caseId", "名称", "级别","步骤", "预期结果")
    # 分别确定测试用例起始行数，以及名称、级别、步骤、预期结果所在的列数
    for row_index, row in enumerate(sheet.iter_rows(min_row=1, max_row=sheet.max_row, values_only=True), start=1):
        flag = 0
        for element in elements:
            if any(x == element for x in row):
                flag += 1
                if element == "名称":
                    caseNameCol = row.index(element)
                if element == "级别":
                    caseLevelCol = row.index(element)
                if element == "步骤":
                    caseStepCol = row.index(element)
                if element == "预期结果":
                    caseResultCol = row.index(element)
                if element == "前提条件":
                    casePreconditionCol = row.index(element)
        if flag == 5:
            caseStartRow = row_index
            break
    # 根据行数、列数将测试用例组合到一个list中
    dcaseContent = []
    for row in sheet.iter_rows(min_row=caseStartRow + 1, max_row=sheet.max_row, values_only=True):
        dcaseContent.append((row[caseNameCol], row[caseStepCol],
                             row[caseResultCol], row[caseLevelCol],
                             row[casePreconditionCol]))
    wb.close()
    return dcaseContent


def formatListContent(tempList, keyStr):
    """
    根据传入的keyStr，按照oe能识别的字符串格式进行拼接
    :param tempList:
    :param keyStr:
    :return:
    """
    tempStr = ""
    for i in range(len(tempList)):
        tempStr += f"{keyStr}{i+1}:{tempList[i]}\n"
    return tempStr


def deleteSpace(tempList):
    """
    删除list中包含的空串
    :param tempList:
    :return:
    """
    newList = [x for x in tempList if x is not None and str(x).strip() != "" and not (isinstance(x, list) and not x)]
    return newList


def parseCase(dcaseContent):
    """
    将测试步骤、验证结果以及case等级修改为oe所能识别的样式，并原格式返回
    :param dcaseContent:
    :return:
    """
    tempConent =[]
    for one in dcaseContent:
        try:
            tempStep = formatListContent(deleteSpace(one[1].split("\n")), "step")
            tempResult = formatListContent(deleteSpace(one[2].split("\n")), "assert")
            tempLevel = one[3].replace("D", "p")
            tempConent.append((one[0], tempStep, tempResult,tempLevel, one[4]))
        except Exception as e:
            logging.error(e)
    return tempConent


def createOECase(dcaseContent):
    """
        1、复制oe case模板，创新一个新的工作表
        2、判断用例名、优先级、步骤、校验点、前置条件在模板中所在的列
        3、将从Dcase中解析获取的数据写入到新的工作表中进行输出
        :param dcaseContent:
        :return:
        """
    # 复制模板。从oe下载的模板中有隐藏列会导致打开失败，所以此脚本的模板将隐藏列取消了隐藏
    outCaseDir = os.path.join(CURRENT_DIR, "out.xlsx")
    shutil.copyfile(OE_EXCEL_TEMPLATE_DIR, outCaseDir)

    wb = openpyxl.load_workbook(outCaseDir)
    sheet = wb["Data"]
    oeCaseNameCol = 0
    oeCaseLevelCol = 0
    oeCaseStepCol = 0
    oeCaseResultCol = 0
    oeCasePreconditionCol = 0
    for cell in sheet[1]:
        if cell.value == "用例名":
            oeCaseNameCol = cell.column
        if cell.value == "优先级":
            oeCaseLevelCol = cell.column
        if cell.value == "步骤":
            oeCaseStepCol = cell.column
        if cell.value == "校验点":
            oeCaseResultCol = cell.column
        if cell.value == "前置条件":
            oeCasePreconditionCol = cell.column

    for i in range(len(dcaseContent)):
        oneCase = dcaseContent[i]
        sheet.cell(row=i + 2, column=oeCaseNameCol, value=oneCase[0])
        sheet.cell(row=i + 2, column=oeCaseStepCol,value=oneCase[1])
        sheet.cell(row=i + 2, column=oeCaseResultCol, value=oneCase[2])
        sheet.cell(row=i + 2, column=oeCaseLevelCol, value=oneCase[3])
        sheet.cell(row=i + 2, column=oeCasePreconditionCol, value=oneCase[4])
    wb.save(outCaseDir)


# def save_coverage():
#     cov.stop()
#     cov.save()
#     cov.html_report()


if __name__ == "__main__":
    try:
        # dcaseDir: str = "/Users/didi/Downloads/模板case转.xlsx"
        # dcaseContent = getDcaseData(dcaseDir)
        dcaseContent = getDcaseData(sys.argv[1])
        createOECase(parseCase(dcaseContent))
    except Exception as e:
        logging.error(e)

# atexit.register(save_coverage)