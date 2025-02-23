import openpyxl
import os
import shutil
import sys
import logging


parent_folder_path = os.path.dirname(os.path.dirname(__file__))
print(f'parent_folder_path{parent_folder_path}')
oe_path = parent_folder_path + '/data'
OE_EXCEL_TEMPLATE_DIR = os.path.join(oe_path, "import_excel_template_27772.xlsx")

def getDcaseData(dcaseDir):
    try:
        wb = openpyxl.load_workbook(dcaseDir)
        sheet = wb.worksheets[0]
        caseStartRow = 0
        caseNameCol = 0
        caseLevelCol = 0
        caseStepCol = 0
        caseResultCol = 0
        casePreconditionCol = 0
        elements = ("caseId", "名称", "级别", "步骤", "预期结果")
        for row_index, row in enumerate(sheet.iter_rows(min_row=1, max_row=sheet.max_row, values_only=True), start=1):
            flag = 0
            for element in elements:
                if any(x == element for x in row):
                    flag += 1
                    if element == "名称":
                        caseNameCol = row.index(element)
                    if element == "级别":
                        caseLevelCol = row.index(element)
                    if element == "步骤":
                        caseStepCol = row.index(element)
                    if element == "预期结果":
                        caseResultCol = row.index(element)
                    if element == "前提条件":
                        casePreconditionCol = row.index(element)
            if flag == 5:
                caseStartRow = row_index
                break
        dcaseContent = []
        for row in sheet.iter_rows(min_row=caseStartRow + 1, max_row=sheet.max_row, values_only=True):
            dcaseContent.append((row[caseNameCol], row[caseStepCol],
                                 row[caseResultCol], row[caseLevelCol],
                                 row[casePreconditionCol]))
        wb.close()
        return dcaseContent
    except Exception as e:
        logging.error(f"解析 Dcase 测试用例文件 {dcaseDir} 时出错: {e}")
        return []

def formatListContent(tempList, keyStr):
    tempStr = ""
    for i in range(len(tempList)):
        tempStr += f"{keyStr}{i + 1}:{tempList[i]}\n"
    return tempStr

def deleteSpace(tempList):
    newList = [x for x in tempList if x is not None and str(x).strip() != "" and not (isinstance(x, list) and not x)]
    return newList

def parseCase(dcaseContent):
    tempConent = []
    for index, one in enumerate(dcaseContent):
        try:
            tempStep = formatListContent(deleteSpace(one[1].split("\n")), "step")
            tempResult = formatListContent(deleteSpace(one[2].split("\n")), "assert")
            tempLevel = one[3].replace("D", "p")
            tempConent.append((one[0], tempStep, tempResult, tempLevel, one[4]))
        except Exception as e:
            logging.error(f"处理第 {index + 1} 条测试用例时出错: {e}")
    return tempConent

def createOECase(dcaseContent):
    try:
        outCaseDir = os.path.join(oe_path, "out.xlsx")
        shutil.copyfile(OE_EXCEL_TEMPLATE_DIR, outCaseDir)
        wb = openpyxl.load_workbook(outCaseDir)
        sheet = wb["Data"]
        oeCaseNameCol = 0
        oeCaseLevelCol = 0
        oeCaseStepCol = 0
        oeCaseResultCol = 0
        oeCasePreconditionCol = 0
        for cell in sheet[1]:
            if cell.value == "用例名":
                oeCaseNameCol = cell.column
            if cell.value == "优先级":
                oeCaseLevelCol = cell.column
            if cell.value == "步骤":
                oeCaseStepCol = cell.column
            if cell.value == "校验点":
                oeCaseResultCol = cell.column
            if cell.value == "前置条件":
                oeCasePreconditionCol = cell.column
        for i in range(len(dcaseContent)):
            oneCase = dcaseContent[i]
            sheet.cell(row=i + 2, column=oeCaseNameCol, value=oneCase[0])
            sheet.cell(row=i + 2, column=oeCaseStepCol, value=oneCase[1])
            sheet.cell(row=i + 2, column=oeCaseResultCol, value=oneCase[2])
            sheet.cell(row=i + 2, column=oeCaseLevelCol, value=oneCase[3])
            sheet.cell(row=i + 2, column=oeCasePreconditionCol, value=oneCase[4])
        wb.save(outCaseDir)
    except Exception as e:
        logging.error(f"创建 OE 测试用例文件时出错: {e}")

if __name__ == "__main__":
    try:
        if len(sys.argv) < 2:
            logging.error("请提供 Dcase 测试用例 Excel 文件的路径作为命令行参数。")
        else:
            dcaseContent = getDcaseData(sys.argv[1])
            createOECase(parseCase(dcaseContent))
    except Exception as e:
        logging.error(f"主程序运行时出错: {e}")